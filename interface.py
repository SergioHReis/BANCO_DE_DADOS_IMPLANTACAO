import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from openpyxl import load_workbook
from datetime import datetime
from dateutil.parser import parse
from dateutil.relativedelta import relativedelta
from tkinter import messagebox
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from PIL import Image, ImageTk

class ControleInformacoes:
    def __init__(self, root):
        self.root = root
        self.root.title("Controle de Informações e Prazos - Criado por Sérgio Henrique Reis Sá")

        self.informacoes = []

        main_frame = ttk.Frame(root)
        main_frame.pack(padx=10, pady=10, fill="both", expand=True)

        # Frame for the buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(side="left", padx=10, fill="y")

        # Frame for the chart and table
        chart_and_table_frame = ttk.Frame(main_frame)
        chart_and_table_frame.pack(side="right", fill="both", expand=True)

        self.inputs = {}

        buscar_arquivo_button = ttk.Button(button_frame, text="Buscar Arquivo Excel", command=self.buscar_arquivo_excel)
        buscar_arquivo_button.grid(row=0, column=0, pady=10, sticky="ew")

        status_atual_button = ttk.Button(button_frame, text="Status Atual", command=self.mostrar_status_atual)
        status_atual_button.grid(row=1, column=0, pady=10, sticky="ew")

        prazo_dc_button = ttk.Button(button_frame, text="Prazo DC's", command=self.calcular_prazos)
        prazo_dc_button.grid(row=2, column=0, pady=10, sticky="ew")

        self.status_atual_info = tk.StringVar()
        label_status_atual_info = ttk.Label(button_frame, textvariable=self.status_atual_info)
        label_status_atual_info.grid(row=3, column=0, pady=10, sticky="w")

        self.total_dc_dentro_do_prazo = tk.StringVar()
        self.total_dc_fora_do_prazo = tk.StringVar()

        label_dc_dentro_do_prazo = ttk.Label(button_frame, textvariable=self.total_dc_dentro_do_prazo, foreground="green")
        label_dc_dentro_do_prazo.grid(row=4, column=0, padx=10, pady=10, sticky="w")

        label_dc_fora_do_prazo = ttk.Label(button_frame, textvariable=self.total_dc_fora_do_prazo, foreground="red")
        label_dc_fora_do_prazo.grid(row=5, column=0, padx=10, pady=10, sticky="w")

        self.b2b_button = ttk.Button(button_frame, text="Contar Projetos B2B", command=self.contar_projetos_b2b)
        self.b2b_button.grid(row=6, column=0, pady=10, sticky="ew")
        self.quantidade_projetos_b2b = tk.StringVar()
        label_quantidade_b2b = ttk.Label(button_frame, textvariable=self.quantidade_projetos_b2b)
        label_quantidade_b2b.grid(row=7, column=0, padx=10, pady=10, sticky="w")

        self.colunas_excel = []

        # Frame for the chart
        graph_frame = ttk.Frame(chart_and_table_frame)
        graph_frame.pack(side="top", fill="both", expand=True)

        self.fig, self.ax = plt.subplots()
        self.ax.axis("off")
        self.canvas = FigureCanvasTkAgg(self.fig, master=graph_frame)
        self.canvas_widget = self.canvas.get_tk_widget()
        self.canvas_widget.pack(padx=10, pady=10, fill="both", expand=True, anchor="nw")

        carregar_imagem_button = ttk.Button(graph_frame, text="Carregar Imagem de Fundo", command=self.carregar_imagem_de_fundo)
        carregar_imagem_button.pack(pady=10, anchor="w")

        self.label_imagem_fundo = ttk.Label(graph_frame)
        self.label_imagem_fundo.pack(padx=10, pady=10)

        # Frame for the table
        table_frame = ttk.Frame(chart_and_table_frame)
        table_frame.pack(side="bottom", fill="both", expand=True)

        self.table_tree = ttk.Treeview(table_frame, columns=["Dias de Atraso", "Quantidade"], show="headings")
        self.table_tree.heading("Dias de Atraso", text="DIAS")
        self.table_tree.heading("Quantidade", text="Obras")
        self.table_tree.column("Dias de Atraso", width=100)
        self.table_tree.column("Quantidade", width=100)
        self.table_tree.pack(padx=10, pady=10, fill="both", expand=True)

    def buscar_arquivo_excel(self):
        arquivo_excel = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx")])
        if arquivo_excel:
            self.carregar_dados_do_excel(arquivo_excel)

    def carregar_dados_do_excel(self, arquivo_excel):
        try:
            workbook = load_workbook(arquivo_excel)
            planilha = workbook.active

            self.colunas_excel = [cell.value for cell in planilha[1]]

            for row in planilha.iter_rows(min_row=2, values_only=True):
                dados = dict(zip(self.colunas_excel, row))
                self.informacoes.append(dados)

            self.contar_projetos_b2b()
            self.criar_grafico_pizza()
            print("Dados do Excel carregados com sucesso!")
        except Exception as e:
            print(f"Erro ao carregar dados do Excel: {e}")

    def calcular_prazos(self):
        hoje = datetime.now()

        total_dc_dentro_do_prazo = 0
        total_dc_fora_do_prazo = 0
        dias_em_atraso = {}
        categorias = ["Em Execução", "Programação da Execução", "Serviço Liberado para Execução"]

        for info in self.informacoes:
            data_status_text = info.get("Data Status Atual")
            if isinstance(data_status_text, datetime):
                data_status_text = data_status_text.strftime("%Y-%m-%d")
                info["Data Status Atual"] = data_status_text
            if data_status_text:
                try:
                    data_status = datetime.strptime(data_status_text, "%Y-%m-%d")
                    prazo = relativedelta(hoje, data_status)
                    dias_em_atraso[prazo.days] = dias_em_atraso.get(prazo.days, 0) + 1

                    categoria = info.get("Status Atual", "").strip()
                    if any(cat in categoria for cat in categorias):
                        if prazo.days > 10:
                            info["Prazo"] = "DC's FORA DO PRAZO"
                            total_dc_fora_do_prazo += 1
                        else:
                            info["Prazo"] = "DC's DENTRO DO PRAZO"
                            total_dc_dentro_do_prazo += 1
                    else:
                        info["Prazo"] = ""
                except ValueError:
                    messagebox.showerror("Erro", "Data inválida na coluna 'Data Status Atual' para uma das entradas.")
                    return

        dias_em_atraso = dict(sorted(dias_em_atraso.items()))

        self.atualizar_labels_totais(total_dc_dentro_do_prazo, total_dc_fora_do_prazo)
        self.mostrar_tabela_dias_atraso(dias_em_atraso)

    def atualizar_labels_totais(self, total_dc_dentro_do_prazo, total_dc_fora_do_prazo):
        self.total_dc_dentro_do_prazo.set(f"DC's dentro do Prazo ({total_dc_dentro_do_prazo})")
        self.total_dc_fora_do_prazo.set(f"DCs Fora do Prazo ({total_dc_fora_do_prazo})")

    def contar_projetos_b2b(self):
        total_projetos_b2b = 0
        for info in self.informacoes:
            if info.get("Tipo de Projeto") == "Projeto Ultra Acelerado - B2B":
                total_projetos_b2b += 1
        self.quantidade_projetos_b2b.set(f"Projetos Ultra Acelerado - B2B: {total_projetos_b2b}")

    def mostrar_status_atual(self):
        servico_paralisado = 0
        em_execucao = 0
        pendencia_cliente = 0
        programacao_execucao = 0
        servico_liberado_execucao = 0

        for info in self.informacoes:
            status_atual = info.get("Status Atual", "").strip().lower()

            if "paralisado" in status_atual:
                servico_paralisado += 1
            elif "execução" in status_atual:
                em_execucao += 1
            elif "pendência" in status_atual:
                pendencia_cliente += 1
            elif "programação" in status_atual:
                programacao_execucao += 1
            elif "liberado para execução" in status_atual:
                servico_liberado_execucao += 1

        status_info = (
            f"Serviço Paralisado: {servico_paralisado}\n"
            f"Em Execução: {em_execucao}\n"
            f"Pendência Cliente: {pendencia_cliente}\n"
            f"Programação da Execução: {programacao_execucao}\n"
            f"Serviço Liberado para Execução: {servico_liberado_execucao}"
        )
        self.status_atual_info.set(status_info)

    def mostrar_tabela_dias_atraso(self, dias_em_atraso):
        for item in self.table_tree.get_children():
            self.table_tree.delete(item)

        for dia, quantidade in dias_em_atraso.items():
            self.table_tree.insert("", "end", values=[dia, quantidade])

    def criar_grafico_pizza(self):
        status_counts = {}
        for info in self.informacoes:
            status_atual = info.get("Status Atual", "").strip()
            if status_atual:
                status_counts[status_atual] = status_counts.get(status_atual, 0) + 1

        labels = status_counts.keys()
        sizes = status_counts.values()

        self.ax.clear()
        self.ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140)
        self.ax.axis('equal')

        self.canvas.draw()

    def carregar_imagem_de_fundo(self):
        imagem_fundo_path = filedialog.askopenfilename(filetypes=[("Arquivos de Imagem", "*.png *.jpg *.jpeg *.gif")])
        if imagem_fundo_path:
            imagem_fundo = Image.open(imagem_fundo_path)
            imagem_fundo = imagem_fundo.resize((800, 600), Image.ANTIALIAS)
            imagem_fundo = ImageTk.PhotoImage(imagem_fundo)
            self.label_imagem_fundo.config(image=imagem_fundo)
            self.label_imagem_fundo.image = imagem_fundo
            self.fig.figimage(imagem_fundo, 100, 100)

    def mostrar_tabela_prazos(self, prazos):
        for item in self.table_tree.get_children():
            self.table_tree.delete(item)

        for prazo, status in prazos.items():
            self.table_tree.insert("", "end", values=[prazo, status])

    def atualizar_treeview(self):
        for item in self.table_tree.get_children():
            self.table_tree.delete(item)

        for info in self.informacoes:
            data_status = info.get("Data Status Atual")
            prazo = info.get("Prazo")
            if data_status and prazo:
                self.table_tree.insert("", "end", values=[data_status, prazo])

if __name__ == "__main__":
    root = tk.Tk()
    app = ControleInformacoes(root)
    root.geometry("1200x800")
    root.mainloop()
